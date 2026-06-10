# -*- coding: utf-8 -*-
"""
Génération du classeur « Tableau MMS » — tableaux + graphiques côte à côte par feuille.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from config.hypotheses_marche import charger_hypotheses
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (clé PNG, titre, clé explication)
_BLOCS_PAR_FEUILLE: dict[str, list[tuple[str, str, str]]] = {
    "Historique": [
        ("interventions_journalieres", "Activité quotidienne", "courbe_journaliere"),
        ("tranches_delai", "Répartition des délais", "tranches_delai"),
    ],
    "Synthèse par appareil": [
        ("pannes_par_appareil", "Pannes par appareil", "pannes_appareil"),
        ("top10_penalites", "Top 10 exposition", "top10"),
    ],
    "Excès pannes": [
        ("pannes_cumul_periode", "Cumul pannes période", "exces_pannes"),
    ],
    "Maintenance": [
        ("delta_maintenance", "Écart visites (Δ jours)", "delta_maint"),
        ("heatmap_maintenance", "Calendrier visites", "heatmap_maint"),
    ],
    "Parachute & Câbles": [],
    "Pénalités": [
        ("penalites_par_type", "Répartition pénalités (€)", "penalites_type"),
    ],
}

_LARGEUR_IMAGE = 400
_HAUTEUR_LIGNE_IMAGE = 14
_COLONNES_ZONE_GRAPHIQUE = 10


def _ajuster_largeurs_tableau(ws, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 28)


def _style_entete_tableau(ws, row: int, max_col: int) -> None:
    fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _surligner_exces_pannes(ws, df: pd.DataFrame, seuil_mensuel: float = 0) -> None:
    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    if df.empty:
        return
    id_cols = {"cle_appareil", "ascenseur", "client", "adresse"}
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, col_name in enumerate(df.columns, start=1):
            if col_name in id_cols:
                continue
            val = getattr(row, col_name, None) if hasattr(row, col_name) else row[c_idx - 1]
            try:
                if float(val) > seuil_mensuel:
                    ws.cell(row=r_idx, column=c_idx).fill = fill
            except (TypeError, ValueError):
                pass


def _ecrire_feuille(writer: pd.ExcelWriter, nom: str, df: pd.DataFrame) -> None:
    if df is None or df.empty:
        pd.DataFrame({"Info": ["Aucune donnée pour cette section"]}).to_excel(
            writer, sheet_name=nom[:31], index=False
        )
    else:
        df.to_excel(writer, sheet_name=nom[:31], index=False)


def _construire_explications(
    resultats: dict[str, Any],
    params: dict[str, Any],
    penalites: dict[str, Any],
) -> dict[str, str]:
    pen = penalites or {}
    detail = pen.get("detail", {}) or {}
    ind = resultats.get("indicateurs") if isinstance(resultats.get("indicateurs"), dict) else {}
    interventions = resultats.get("interventions", pd.DataFrame())
    nb_inter = len(interventions) if isinstance(interventions, pd.DataFrame) else 0
    trimestre = params.get("trimestre", "T1")
    seuil_delai = params.get("seuil_delai_h", 2)
    seuil_maint = params.get("seuil_maintenance_jours", 42)
    client = params.get("filtre_client") or params.get("hypotheses_client") or "Tous"
    prest = params.get("prestataire", "")

    tranches = ind.get("tranches_delai") or {}
    if not tranches and isinstance(interventions, pd.DataFrame) and "tranche_delai" in interventions.columns:
        tranches = interventions["tranche_delai"].value_counts().to_dict()

    nb_gt4 = int(tranches.get("> 4h", 0))
    nb_inconnu = int(tranches.get("inconnu", 0))
    pct_gt4 = round(100 * nb_gt4 / nb_inter, 1) if nb_inter else 0

    return {
        "courbe_journaliere": (
            f"Période {trimestre} {params.get('annee', '')} — {client} / {prest}.\n"
            f"Volume d'interventions jour par jour ; repérer les pics d'activité.\n"
            f"Un pic isolé peut masquer plusieurs retards > {seuil_delai} h le même jour."
        ),
        "tranches_delai": (
            f"{nb_inter} interventions classées par délai d'arrivée technicien.\n"
            f"Vert = < 2 h ; orange = 2–4 h ; rouge = > 4 h (seuil marché : {seuil_delai} h).\n"
            f"{nb_gt4} cas > 4 h ({pct_gt4} %) ; {nb_inconnu} horaires non exploitables."
        ),
        "pannes_appareil": (
            "Pannes « retenues » par appareil selon critères contractuels.\n"
            "Barres hautes = priorité de suivi avec le prestataire."
        ),
        "top10": (
            "Classement composite (indisponibilité + pannes × 10).\n"
            "Ces appareils concentrent le risque pénalités sur le trimestre."
        ),
        "exces_pannes": (
            "Somme des pannes retenues sur la période (matrice mensuelle).\n"
            "Surveiller les appareils en tête pour le bilan annuel (seuil 2/an)."
        ),
        "delta_maint": (
            f"Δ jours entre visites consécutives — seuil {seuil_maint} j.\n"
            f"Rouge : écart contractuel ; vert : conforme ; gris : 1ʳᵉ visite sans historique T-1."
        ),
        "heatmap_maint": (
            "Une case = au moins une visite dans la semaine ISO.\n"
            "Trous vert pâle = absence de passage maintenance."
        ),
        "penalites_type": (
            f"Total {pen.get('penalite_totale', 0):,.0f} € — ventilation par poste contractuel.\n"
            f"Maintenance {pen.get('penalite_maintenance', 0):,.0f} € | "
            f"Pannes {pen.get('penalite_pannes', 0):,.0f} € | "
            f"Délai {pen.get('penalite_delai_intervention', 0):,.0f} €."
        ),
    }


def _inserer_image(ws, chemin: str, cellule: str, largeur: int = _LARGEUR_IMAGE) -> int:
    p = Path(chemin)
    if not p.is_file():
        return 0
    try:
        img = XLImage(str(p))
        ratio = largeur / img.width if img.width > largeur else 1.0
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
        ws.add_image(img, cellule)
        return max(12, int(img.height / _HAUTEUR_LIGNE_IMAGE) + 1)
    except OSError:
        return 0


def _preparer_zone_graphique(ws, col_debut: int) -> None:
    """Largeur fixe pour la colonne des graphiques à droite du tableau."""
    for offset in range(_COLONNES_ZONE_GRAPHIQUE):
        ws.column_dimensions[get_column_letter(col_debut + offset)].width = 11


def _integrer_bloc_a_cote(
    ws,
    row_depart: int,
    col_depart: int,
    chemin: Optional[str],
    titre: str,
    explication: str,
) -> int:
    """Graphique + légende à droite du tableau. Retourne la prochaine ligne disponible."""
    if not chemin:
        return row_depart

    border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    fill_titre = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

    ws.cell(row=row_depart, column=col_depart, value=titre)
    c_titre = ws.cell(row=row_depart, column=col_depart)
    c_titre.font = Font(bold=True, size=10, color="FFFFFF")
    c_titre.fill = fill_titre
    c_titre.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=row_depart,
        start_column=col_depart,
        end_row=row_depart,
        end_column=col_depart + min(4, _COLONNES_ZONE_GRAPHIQUE - 1),
    )

    row = row_depart + 1
    lignes = [ln.strip() for ln in explication.split("\n") if ln.strip()]
    for i, line in enumerate(lignes[:4]):
        cell = ws.cell(row=row + i, column=col_depart, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=9, color="1F3A5F")
        ws.row_dimensions[row + i].height = 32

    row += len(lignes[:4]) + 1
    cellule_img = f"{get_column_letter(col_depart)}{row}"
    h_img = _inserer_image(ws, chemin, cellule_img)
    return row + h_img + 2


def _integrer_note_parachute(ws, col_graph: int, params: dict[str, Any]) -> None:
    ws.insert_rows(1, 2)
    ws["A1"] = "Parachute & Câbles — suivi réglementaire"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    note = (
        f"Tests parachute (seuil {params.get('seuil_parachute_mois', 12)} mois) "
        f"et câbles ({params.get('seuil_cables_mois', 6)} mois).\n"
        "Tableau à gauche : dates de contrôle par appareil."
    )
    c = ws.cell(row=3, column=col_graph, value=note)
    c.alignment = Alignment(wrap_text=True)
    c.font = Font(size=9, color="1F3A5F")


def _integrer_analyses_par_feuille(
    wb,
    figure_paths: dict[str, str],
    explications: dict[str, str],
    params: Optional[dict[str, Any]] = None,
) -> None:
    """Place chaque graphique à droite du tableau de la feuille correspondante."""
    row_table_header = 3

    for feuille, blocs in _BLOCS_PAR_FEUILLE.items():
        if feuille not in wb.sheetnames:
            continue
        ws = wb[feuille]
        blocs_ok = [(cle, titre, tk) for cle, titre, tk in blocs if cle in figure_paths]

        if feuille == "Parachute & Câbles" and not blocs_ok:
            data_cols = max(1, ws.max_column)
            _integrer_note_parachute(ws, data_cols + 2, params or {})
            continue
        if not blocs_ok:
            continue

        data_cols = max(1, ws.max_column)
        col_graph = data_cols + 2

        ws.insert_rows(1, 2)
        ws["A1"] = f"{feuille}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=data_cols)
        ws["A2"] = "← Données détaillées          |          Synthèse graphique →"
        ws["A2"].font = Font(italic=True, size=9, color="5D6D7E")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_graph + 4)

        data_cols = max(1, ws.max_column)
        col_graph = data_cols + 2
        _preparer_zone_graphique(ws, col_graph)
        _style_entete_tableau(ws, row_table_header, data_cols)

        row_cur = row_table_header
        for cle, titre, txt_key in blocs_ok:
            exp = explications.get(txt_key, "")
            row_cur = _integrer_bloc_a_cote(
                ws,
                row_cur,
                col_graph,
                figure_paths.get(cle),
                titre,
                exp,
            )

        _ajuster_largeurs_tableau(ws, data_cols)
        ws.freeze_panes = "A4"


def _integrer_graphiques_recapitulatif(
    wb, figure_paths: dict[str, str], explications: dict[str, str]
) -> None:
    if not figure_paths:
        return
    if "Graphiques" in wb.sheetnames:
        del wb["Graphiques"]
    ws = wb.create_sheet("Graphiques")
    ws["A1"] = "Index des graphiques — LVO"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")
    row = 3
    txt_keys = {cle: tk for blocs in _BLOCS_PAR_FEUILLE.values() for cle, _, tk in blocs}
    for cle, chemin in figure_paths.items():
        if cle == "carte_penalites":
            continue
        titre = cle.replace("_", " ").title()
        exp = explications.get(txt_keys.get(cle, ""), "")
        row = _integrer_bloc_a_cote(ws, row, 1, chemin, titre, exp)


def generer_tableau_mms(
    resultats: dict[str, Any],
    params: dict[str, Any],
    seuil_pannes_an: int = 2,
    figure_paths: Optional[dict[str, str]] = None,
) -> bytes:
    buffer = io.BytesIO()
    pen = resultats.get("penalites", {})
    explications = _construire_explications(resultats, params, pen)

    interventions = resultats.get("interventions", pd.DataFrame())
    cols_hist = [c for c in interventions.columns if c in list(interventions.columns)]

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _ecrire_feuille(writer, "Historique", interventions[cols_hist] if not interventions.empty else interventions)
        _ecrire_feuille(writer, "Synthèse par appareil", resultats.get("synthese_appareils", pd.DataFrame()))
        _ecrire_feuille(writer, "Excès pannes", resultats.get("matrice_pannes", pd.DataFrame()))
        _ecrire_feuille(writer, "Maintenance", resultats.get("maintenance", pd.DataFrame()))
        _ecrire_feuille(writer, "Parachute & Câbles", resultats.get("suivi_tests", pd.DataFrame()))

        detail = pen.get("detail", {})
        df_pen = pd.DataFrame(
            [
                {"Poste": "Pénalité maintenance", "Montant (€)": pen.get("penalite_maintenance", 0),
                 "Détail": f"{detail.get('jours_retard_maintenance', 0)} j retard"},
                {"Poste": "Pénalité pannes", "Montant (€)": pen.get("penalite_pannes", 0),
                 "Détail": f"{detail.get('nb_pannes_excedentaires', 0)} excès"},
                {"Poste": "Pénalité immobilisation", "Montant (€)": pen.get("penalite_immobilisation", 0),
                 "Détail": f"{detail.get('heures_immo_excedentaires', 0)} h"},
                {"Poste": "Pénalité délai", "Montant (€)": pen.get("penalite_delai_intervention", 0),
                 "Détail": f"{detail.get('heures_delai_excedentaires', 0)} h"},
                {"Poste": "TOTAL", "Montant (€)": pen.get("penalite_totale", 0), "Détail": ""},
            ]
        )
        _ecrire_feuille(writer, "Pénalités", df_pen)

        client_hyp = params.get("hypotheses_client") or params.get("hypotheses_client_applique")
        hyp_rows = [{"Client": cle, **profil} for cle, profil in charger_hypotheses().items()]
        df_hyp = pd.DataFrame(hyp_rows)
        if client_hyp:
            ligne_act = df_hyp[df_hyp["Client"].astype(str).str.upper() == str(client_hyp).upper()]
            if not ligne_act.empty:
                df_hyp = pd.concat(
                    [pd.DataFrame([{"Info": f"Profil : {client_hyp}"}]), ligne_act, df_hyp],
                    ignore_index=True,
                )
        _ecrire_feuille(writer, "Hypothèses marché", df_hyp)

    buffer.seek(0)
    wb = load_workbook(buffer)
    buffer = io.BytesIO()

    if "Excès pannes" in wb.sheetnames:
        _surligner_exces_pannes(wb["Excès pannes"], resultats.get("matrice_pannes", pd.DataFrame()))
    if "Pénalités" in wb.sheetnames:
        for row in wb["Pénalités"].iter_rows(min_row=2, max_row=wb["Pénalités"].max_row):
            if row[0].value == "TOTAL":
                for cell in row:
                    cell.font = Font(bold=True)

    if figure_paths:
        _integrer_analyses_par_feuille(wb, figure_paths, explications, params)
        _integrer_graphiques_recapitulatif(wb, figure_paths, explications)

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
